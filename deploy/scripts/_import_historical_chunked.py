#!/usr/bin/env python3
"""
ARIS 4.0 — Import large Monthly Animal Health Report in chunks
──────────────────────────────────────────────────────────────
Splits the 801K-row workbook into ~100K-row chunks and imports each
via the datalake API on staging/prod.
"""
from __future__ import annotations

import json
import os
import sys
import time
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

import openpyxl
import paramiko

SSH_USER = os.environ.get("ARIS_DEPLOY_USER", "arisadmin")
SSH_PASS = os.environ.get("ARIS_DEPLOY_PASS", "@u-1baR.0rg$U24")

ENV_CONFIG = {
    "prod": {"host": "10.202.101.183"},
    "stg": {"host": "10.202.101.146"},
}

ROOT = Path("C:/new.au-aris.org/DONNEES-HISTORIQUES-ARIS")
SOURCE = ROOT / "mapped" / "_workbooks" / "animal_health__AU-IBAR_Monthly_Animal_Health_Report__filled.xlsx"
CHUNK_DIR = ROOT / "mapped" / "_workbooks" / "_chunks"
CHUNK_SIZE = 100_000


def login(ssh, base_url):
    body = '{"email":"admin@au-aris.org","password":"Aris2026@@4!0"}'
    cmd = f'curl -sk -X POST "{base_url}/api/v1/credential/auth/login" -H "Content-Type: application/json" -d \'{body}\''
    _, out, _ = ssh.exec_command(cmd, timeout=20)
    return json.loads(out.read().decode())["data"]["accessToken"]


def upload(ssh, sftp, base_url, token, local_path, meta):
    file_name = local_path.name
    remote = f"/tmp/aris_chunk_{file_name}"

    size_mb = local_path.stat().st_size / 1024 / 1024
    print(f"  [sftp] uploading {file_name} ({size_mb:.1f} MB)...", flush=True)
    sftp.put(str(local_path), remote)

    tags_json = json.dumps(meta.get("tags", []))
    cmd = (
        f'curl -sk -X POST "{base_url}/api/v1/historical/import" '
        f'-H "Authorization: Bearer {token}" '
        f'-F "file=@{remote};filename={file_name}" '
        f'-F "name={meta["name"]}" '
        f'-F "domain={meta["domain"]}" '
        f'-F "description={meta.get("description", "")}" '
        f"-F 'tags={tags_json}' "
        f'--max-time 600'
    )
    print(f"  [api] importing...", flush=True)
    _, out, _ = ssh.exec_command(cmd, timeout=660)
    raw = out.read().decode("utf-8", errors="replace")

    try:
        sftp.remove(remote)
    except Exception:
        pass

    try:
        resp = json.loads(raw)
        if "data" in resp:
            ds = resp["data"]
            print(f"  [ok] id={ds.get('id','?')}, rows={ds.get('rowCount','?')}, status={ds.get('status','?')}", flush=True)
            return ds
        else:
            print(f"  [err] {raw[:300]}", flush=True)
            return None
    except json.JSONDecodeError:
        print(f"  [err] non-JSON: {raw[:300]}", flush=True)
        return None


def split():
    """Split the large workbook into chunks."""
    CHUNK_DIR.mkdir(exist_ok=True)

    print(f"[split] Loading {SOURCE.name}...", flush=True)
    wb = openpyxl.load_workbook(SOURCE, read_only=True)
    ws = wb["Data"]
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    headers = all_rows[0]
    labels = all_rows[1]
    types = all_rows[2]
    data_rows = all_rows[3:]
    total = len(data_rows)
    num_chunks = (total + CHUNK_SIZE - 1) // CHUNK_SIZE

    print(f"[split] {total:,} data rows → {num_chunks} chunks of ≤{CHUNK_SIZE:,}", flush=True)

    chunk_paths = []
    for ci in range(num_chunks):
        start = ci * CHUNK_SIZE
        end = min(start + CHUNK_SIZE, total)
        chunk_rows = data_rows[start:end]
        part = ci + 1

        chunk_name = f"monthly_health_part{part:02d}.xlsx"
        chunk_path = CHUNK_DIR / chunk_name

        if chunk_path.exists():
            size = chunk_path.stat().st_size
            if size > 1000:
                print(f"  Part {part}/{num_chunks}: {len(chunk_rows):,} rows — already exists ({size/1024/1024:.1f} MB)", flush=True)
                chunk_paths.append((chunk_path, part, num_chunks, start, end, len(chunk_rows)))
                continue

        print(f"  Part {part}/{num_chunks}: writing {len(chunk_rows):,} rows...", flush=True)
        cwb = openpyxl.Workbook()
        cws = cwb.active
        cws.title = "Data"
        cws.append(list(headers))
        cws.append(list(labels))
        cws.append(list(types))
        for row in chunk_rows:
            cws.append(list(row))
        cwb.save(chunk_path)
        cwb.close()

        size = chunk_path.stat().st_size
        print(f"    → {chunk_name} ({size/1024/1024:.1f} MB)", flush=True)
        chunk_paths.append((chunk_path, part, num_chunks, start, end, len(chunk_rows)))

    return chunk_paths


def main():
    import argparse
    ap = argparse.ArgumentParser()
    ap.add_argument("--env", default="stg", choices=["prod", "stg"])
    ap.add_argument("--from-part", type=int, default=1, help="Resume from this part number")
    ap.add_argument("--split-only", action="store_true")
    args = ap.parse_args()

    # Step 1: Split
    chunks = split()
    if args.split_only:
        print("\n[split-only] Done.")
        return

    # Step 2: Upload
    cfg = ENV_CONFIG[args.env]
    host = cfg["host"]
    base_url = "https://localhost"

    print(f"\n[ssh] connecting to {host}...", flush=True)
    ssh = paramiko.SSHClient()
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    ssh.connect(host, username=SSH_USER, password=SSH_PASS,
                timeout=20, allow_agent=False, look_for_keys=False)
    sftp = ssh.open_sftp()

    print("[auth] logging in...", flush=True)
    token = login(ssh, base_url)
    print("[auth] ok\n", flush=True)

    results = []
    for chunk_path, part, num_chunks, start, end, count in chunks:
        if part < args.from_part:
            print(f"[{part}/{num_chunks}] skipped (--from-part={args.from_part})")
            continue

        meta = {
            "name": f"Monthly Animal Health Reports (Part {part}/{num_chunks})",
            "domain": "animal_health",
            "description": (
                f"Monthly disease surveillance reports from 46 African countries (2008-2025). "
                f"Part {part}/{num_chunks}, rows {start+1:,}–{end:,} ({count:,} rows)."
            ),
            "tags": ["historical", "disease-surveillance", "outbreaks", f"part-{part}"],
        }

        print(f"\n[{part}/{num_chunks}] === rows {start+1:,}–{end:,} ({count:,} rows) ===", flush=True)
        ds = upload(ssh, sftp, base_url, token, chunk_path, meta)
        results.append((part, ds))
        time.sleep(2)

    sftp.close()
    ssh.close()

    # Summary
    print("\n" + "=" * 60)
    total_rows = 0
    ok = 0
    for part, ds in results:
        if ds and ds.get("status") == "READY":
            ok += 1
            rows = ds.get("rowCount", 0)
            total_rows += rows
            print(f"  Part {part}: {rows:,} rows")
        else:
            print(f"  Part {part}: FAILED")
    print(f"\n{ok}/{len(results)} parts imported — {total_rows:,} total rows")


if __name__ == "__main__":
    main()
