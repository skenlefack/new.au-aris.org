#!/usr/bin/env python3
"""
Seed PAID reference data from Kobo Excel into staging/prod DB.
Creates tables in animal_health schema (shared with master-data service).
"""

import openpyxl
import paramiko
import sys
import json
import os
import time

sys.stdout.reconfigure(encoding='utf-8', errors='replace')

EXCEL = os.path.join(os.path.dirname(__file__), '..', '..', 'form', 'LICS', 'AU_IBAR_PAID_Kobo_Formatted_Repeat_2026.xlsx')

SSH_USER = "arisadmin"
SSH_PASS = "@u-1baR.0rg$U24"

TARGETS = {
    "STAGING": {"db_host": "10.202.101.148", "container": "aris-stg-postgres", "db_pass": "Ar1s_Stg_2024!xK9mZ"},
    "PROD":    {"db_host": "10.202.101.185", "container": "aris-postgres",     "db_pass": "Ar1s_Pr0d_2024!xK9mZ"},
}


def extract_choices(wb, list_name):
    """Extract choices from the 'choices' sheet for a given list_name."""
    ws = wb['choices']
    headers = [str(ws.cell(1, c).value or '').strip() for c in range(1, ws.max_column + 1)]
    items = []
    for r in range(2, ws.max_row + 1):
        ln = str(ws.cell(r, 1).value or '').strip()
        if ln != list_name:
            continue
        row = {}
        for c, h in enumerate(headers, 1):
            val = ws.cell(r, c).value
            if val is not None:
                row[h] = str(val).strip()
        items.append(row)
    return items


def escape_sql(s):
    """Escape single quotes for SQL."""
    if s is None:
        return 'NULL'
    return "'" + str(s).replace("'", "''") + "'"


def build_sql(wb):
    """Build all CREATE TABLE + INSERT statements."""
    sql_parts = []

    # ── 1. PAID Sectors ──
    sectors = extract_choices(wb, 'sector')
    sql_parts.append("""
-- ══════════════════════════════════════════════════════
-- PAID Sectors (7)
-- ══════════════════════════════════════════════════════
DROP TABLE IF EXISTS animal_health.paid_sectors CASCADE;
CREATE TABLE animal_health.paid_sectors (
    id SERIAL PRIMARY KEY,
    code VARCHAR(20) NOT NULL UNIQUE,
    name_en VARCHAR(200) NOT NULL,
    name_fr VARCHAR(200),
    name_es VARCHAR(200)
);
""")
    for s in sectors:
        sql_parts.append(
            f"INSERT INTO animal_health.paid_sectors (code, name_en, name_fr, name_es) VALUES ("
            f"{escape_sql(s.get('name'))}, {escape_sql(s.get('label::English'))}, "
            f"{escape_sql(s.get('label::French'))}, {escape_sql(s.get('label::Spanish'))});"
        )

    # ── 2. PAID Activities (386) ──
    activities = extract_choices(wb, 'activity')
    sql_parts.append("""
-- ══════════════════════════════════════════════════════
-- PAID Activities (386) — with sector filter + unit
-- ══════════════════════════════════════════════════════
DROP TABLE IF EXISTS animal_health.paid_activities CASCADE;
CREATE TABLE animal_health.paid_activities (
    id SERIAL PRIMARY KEY,
    code VARCHAR(20) NOT NULL UNIQUE,
    name_en TEXT NOT NULL,
    name_fr TEXT,
    name_es TEXT,
    sector_code VARCHAR(20),
    unit VARCHAR(100),
    input_category VARCHAR(50),
    tier VARCHAR(10)
);
""")
    for a in activities:
        sql_parts.append(
            f"INSERT INTO animal_health.paid_activities (code, name_en, name_fr, name_es, sector_code, unit, input_category, tier) VALUES ("
            f"{escape_sql(a.get('name'))}, {escape_sql(a.get('label::English'))}, "
            f"{escape_sql(a.get('label::French'))}, {escape_sql(a.get('label::Spanish'))}, "
            f"{escape_sql(a.get('sector'))}, {escape_sql(a.get('unit'))}, "
            f"{escape_sql(a.get('input_category'))}, {escape_sql(a.get('tier'))});"
        )

    # ── 3. PAID Species (40) ──
    species = extract_choices(wb, 'species')
    sql_parts.append("""
-- ══════════════════════════════════════════════════════
-- PAID Species (40) — with sector filter
-- ══════════════════════════════════════════════════════
DROP TABLE IF EXISTS animal_health.paid_species CASCADE;
CREATE TABLE animal_health.paid_species (
    id SERIAL PRIMARY KEY,
    code VARCHAR(20) NOT NULL UNIQUE,
    name_en VARCHAR(200) NOT NULL,
    name_fr VARCHAR(200),
    name_es VARCHAR(200),
    sector_code VARCHAR(20)
);
""")
    for s in species:
        sql_parts.append(
            f"INSERT INTO animal_health.paid_species (code, name_en, name_fr, name_es, sector_code) VALUES ("
            f"{escape_sql(s.get('name'))}, {escape_sql(s.get('label::English'))}, "
            f"{escape_sql(s.get('label::French'))}, {escape_sql(s.get('label::Spanish'))}, "
            f"{escape_sql(s.get('sector'))});"
        )

    # ── 4. PAID Production Systems (117) ──
    prod_sys = extract_choices(wb, 'production_system')
    sql_parts.append("""
-- ══════════════════════════════════════════════════════
-- PAID Production Systems (117) — with sector+species filter
-- ══════════════════════════════════════════════════════
DROP TABLE IF EXISTS animal_health.paid_production_systems CASCADE;
CREATE TABLE animal_health.paid_production_systems (
    id SERIAL PRIMARY KEY,
    code VARCHAR(20) NOT NULL UNIQUE,
    name_en VARCHAR(200) NOT NULL,
    name_fr VARCHAR(200),
    name_es VARCHAR(200),
    sector_code VARCHAR(20),
    species_code VARCHAR(50)
);
""")
    for p in prod_sys:
        sql_parts.append(
            f"INSERT INTO animal_health.paid_production_systems (code, name_en, name_fr, name_es, sector_code, species_code) VALUES ("
            f"{escape_sql(p.get('name'))}, {escape_sql(p.get('label::English'))}, "
            f"{escape_sql(p.get('label::French'))}, {escape_sql(p.get('label::Spanish'))}, "
            f"{escape_sql(p.get('sector'))}, {escape_sql(p.get('species'))});"
        )

    # ── 5. PAID Diseases (162) ──
    diseases = extract_choices(wb, 'disease')
    sql_parts.append("""
-- ══════════════════════════════════════════════════════
-- PAID Diseases (162) — with sector+species filter
-- ══════════════════════════════════════════════════════
DROP TABLE IF EXISTS animal_health.paid_diseases CASCADE;
CREATE TABLE animal_health.paid_diseases (
    id SERIAL PRIMARY KEY,
    code VARCHAR(20) NOT NULL UNIQUE,
    name_en VARCHAR(200) NOT NULL,
    name_fr VARCHAR(200),
    name_es VARCHAR(200),
    sector_code VARCHAR(20),
    species_code VARCHAR(50)
);
""")
    for d in diseases:
        sql_parts.append(
            f"INSERT INTO animal_health.paid_diseases (code, name_en, name_fr, name_es, sector_code, species_code) VALUES ("
            f"{escape_sql(d.get('name'))}, {escape_sql(d.get('label::English'))}, "
            f"{escape_sql(d.get('label::French'))}, {escape_sql(d.get('label::Spanish'))}, "
            f"{escape_sql(d.get('sector'))}, {escape_sql(d.get('species'))});"
        )

    # ── 6. PAID Projects (2915) ──
    projects = extract_choices(wb, 'project')
    sql_parts.append("""
-- ══════════════════════════════════════════════════════
-- PAID Projects (2915) — FAO project symbols
-- ══════════════════════════════════════════════════════
DROP TABLE IF EXISTS animal_health.paid_projects CASCADE;
CREATE TABLE animal_health.paid_projects (
    id SERIAL PRIMARY KEY,
    code VARCHAR(30) NOT NULL UNIQUE,
    title TEXT NOT NULL,
    title_fr TEXT,
    title_es TEXT,
    project_symbol VARCHAR(100),
    country VARCHAR(100)
);
""")
    for p in projects:
        # The 'title' column in choices contains the full project title
        title = p.get('title') or p.get('label::English', '')
        sql_parts.append(
            f"INSERT INTO animal_health.paid_projects (code, title, title_fr, title_es, project_symbol, country) VALUES ("
            f"{escape_sql(p.get('name'))}, {escape_sql(p.get('label::English'))}, "
            f"{escape_sql(p.get('label::French'))}, {escape_sql(p.get('label::Spanish'))}, "
            f"{escape_sql(p.get('project_symbol'))}, {escape_sql(p.get('country'))});"
        )

    # ── 7. PAID Partners International (150) ──
    intl = extract_choices(wb, 'international_partner')
    sql_parts.append("""
-- ══════════════════════════════════════════════════════
-- PAID Partners International (150)
-- ══════════════════════════════════════════════════════
DROP TABLE IF EXISTS animal_health.paid_partners_intl CASCADE;
CREATE TABLE animal_health.paid_partners_intl (
    id SERIAL PRIMARY KEY,
    code VARCHAR(30) NOT NULL UNIQUE,
    name_en VARCHAR(300) NOT NULL,
    name_fr VARCHAR(300),
    name_es VARCHAR(300)
);
""")
    for p in intl:
        sql_parts.append(
            f"INSERT INTO animal_health.paid_partners_intl (code, name_en, name_fr, name_es) VALUES ("
            f"{escape_sql(p.get('name'))}, {escape_sql(p.get('label::English'))}, "
            f"{escape_sql(p.get('label::French'))}, {escape_sql(p.get('label::Spanish'))});"
        )

    # ── 8. PAID Partners National (1186) ──
    national = extract_choices(wb, 'national_partner')
    sql_parts.append("""
-- ══════════════════════════════════════════════════════
-- PAID Partners National (1186) — with country filter
-- ══════════════════════════════════════════════════════
DROP TABLE IF EXISTS animal_health.paid_partners_national CASCADE;
CREATE TABLE animal_health.paid_partners_national (
    id SERIAL PRIMARY KEY,
    code VARCHAR(80) NOT NULL UNIQUE,
    name_en VARCHAR(300) NOT NULL,
    name_fr VARCHAR(300),
    name_es VARCHAR(300),
    country VARCHAR(100)
);
""")
    for p in national:
        sql_parts.append(
            f"INSERT INTO animal_health.paid_partners_national (code, name_en, name_fr, name_es, country) VALUES ("
            f"{escape_sql(p.get('name'))}, {escape_sql(p.get('label::English'))}, "
            f"{escape_sql(p.get('label::French'))}, {escape_sql(p.get('label::Spanish'))}, "
            f"{escape_sql(p.get('country'))});"
        )

    return '\n'.join(sql_parts)


def run_sql_on_target(target_name, sql):
    """Execute SQL on target DB via SSH + docker exec."""
    cfg = TARGETS[target_name]
    print(f"\n{'='*50}")
    print(f"  Seeding PAID ref data on {target_name}")
    print(f"{'='*50}")

    ssh = paramiko.SSHClient()
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    ssh.connect(cfg['db_host'], username=SSH_USER, password=SSH_PASS, timeout=15)

    # Write SQL to temp file on server
    sftp = ssh.open_sftp()
    sftp.open('/tmp/paid_seed.sql', 'w').write(sql)
    sftp.close()

    # Copy into container and execute
    cmds = [
        f"echo '{SSH_PASS}' | sudo -S docker cp /tmp/paid_seed.sql {cfg['container']}:/tmp/paid_seed.sql",
        f"echo '{SSH_PASS}' | sudo -S docker exec -e PGPASSWORD={cfg['db_pass']} {cfg['container']} psql -h localhost -U aris -d aris -f /tmp/paid_seed.sql",
    ]

    for cmd in cmds:
        stdin, stdout, stderr = ssh.exec_command(cmd, timeout=120)
        exit_code = stdout.channel.recv_exit_status()
        out = stdout.read().decode(errors='replace')
        err = stderr.read().decode(errors='replace')

        # Check for errors (ignore password prompts)
        err_lines = [l for l in err.split('\n') if 'ERROR' in l]
        if err_lines:
            for el in err_lines[:5]:
                print(f"  ERROR: {el.strip()}")

    # Verify counts
    verify_sql = """
SELECT 'paid_sectors' as tbl, count(*) FROM animal_health.paid_sectors
UNION ALL SELECT 'paid_activities', count(*) FROM animal_health.paid_activities
UNION ALL SELECT 'paid_species', count(*) FROM animal_health.paid_species
UNION ALL SELECT 'paid_production_systems', count(*) FROM animal_health.paid_production_systems
UNION ALL SELECT 'paid_diseases', count(*) FROM animal_health.paid_diseases
UNION ALL SELECT 'paid_projects', count(*) FROM animal_health.paid_projects
UNION ALL SELECT 'paid_partners_intl', count(*) FROM animal_health.paid_partners_intl
UNION ALL SELECT 'paid_partners_national', count(*) FROM animal_health.paid_partners_national
ORDER BY 1;
"""
    cmd = f"echo '{SSH_PASS}' | sudo -S docker exec -e PGPASSWORD={cfg['db_pass']} {cfg['container']} psql -h localhost -U aris -d aris -c \"{verify_sql}\""
    stdin, stdout, stderr = ssh.exec_command(cmd, timeout=15)
    stdout.channel.recv_exit_status()
    out = stdout.read().decode(errors='replace')
    lines = [l for l in out.split('\n') if l.strip() and 'password' not in l.lower() and not l.startswith('[sudo]')]
    print("\n  Verification:")
    for l in lines:
        print(f"    {l.strip()}")

    ssh.close()
    print(f"\n  OK - {target_name} done.")


if __name__ == '__main__':
    print("Loading Excel...")
    wb = openpyxl.load_workbook(EXCEL, data_only=True)

    print("Building SQL...")
    sql = build_sql(wb)
    print(f"Generated {len(sql)} chars, {sql.count('INSERT')} INSERT statements")

    # Save SQL for review
    sql_file = os.path.join(os.path.dirname(__file__), '_paid_refdata.sql')
    with open(sql_file, 'w', encoding='utf-8') as f:
        f.write(sql)
    print(f"SQL saved to {sql_file}")

    targets = sys.argv[1:] if len(sys.argv) > 1 else ['STAGING']
    for t in targets:
        t = t.upper()
        if t in TARGETS:
            run_sql_on_target(t, sql)
        else:
            print(f"Unknown target: {t}")
