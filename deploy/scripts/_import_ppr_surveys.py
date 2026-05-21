#!/usr/bin/env python3
"""
Phase 3: Import existing PPR survey responses into ARIS submissions.
- Survey A: 41 responses (Surveillance & Digital Tools)
- Survey B: 21 responses (Diagnostic Tests & HPPR-bELISA)
- Dataset C: 21 kit shipments
"""
import paramiko
import json
import uuid
import os
import sys
import hashlib
import tempfile
import zipfile
from xml.etree import ElementTree

sys.stdout.reconfigure(encoding='utf-8', errors='replace')

VM_APP = "10.202.101.183"
SSH_USER = "arisadmin"
SSH_PASS = "@u-1baR.0rg$U24"
DB_HOST = "10.202.101.185"
DB_PASS = "Ar1s_Pr0d_2024!xK9mZ"
TENANT_AU_IBAR = "00000000-0000-4000-a000-000000000001"
USER_SUPER_ADMIN = "10000000-0000-4000-a000-000000000001"

BASE = os.path.dirname(os.path.abspath(__file__))
PPR_DIR = os.path.normpath(os.path.join(BASE, "..", "..", "ppr"))

# Load campaign/template IDs
with open(os.path.join(BASE, "_ppr_survey_ids.json")) as f:
    IDS = json.load(f)

CAMPAIGN_A = IDS["campaigns"]["A"]
CAMPAIGN_B = IDS["campaigns"]["B"]
CAMPAIGN_C = IDS["campaigns"]["C"]
TEMPLATE_A = IDS["templates"]["A"]
TEMPLATE_B = IDS["templates"]["B"]
TEMPLATE_C = IDS["templates"]["C"]


def deterministic_uuid(campaign_id, row_key):
    """Generate a deterministic UUID from campaign+row key for idempotent imports."""
    raw = f"{campaign_id}:{row_key}"
    h = hashlib.sha256(raw.encode()).hexdigest()
    return str(uuid.UUID(h[:32]))


def safe(val):
    """Clean a cell value for SQL insertion."""
    if val is None:
        return None
    s = str(val).strip()
    if s in ('', 'None', 'nan'):
        return None
    return s.replace("'", "''")


# ═══════════════════════════════════════════════
# PARSE SOURCE FILES
# ═══════════════════════════════════════════════

def parse_survey_a():
    """Parse Surveillance survey (46 cols, 41 responses)."""
    from openpyxl import load_workbook
    path = os.path.join(PPR_DIR, "Copy of Systèmes de Surveillance et Outils Numériques en Santé Animale-Responses.xlsx")
    wb = load_workbook(path, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    headers = rows[0]
    submissions = []
    for row in rows[2:]:  # Skip header + label row
        if not row[0]:
            continue
        respondent_id = str(row[0]).strip()
        data = {
            "country": safe(row[10]) or "",  # SurveyMonkey ID (numeric)
            "institution": safe(row[11]) or "",
            "respondent_name": safe(row[12]) or "",
            "respondent_role": safe(row[13]) or "",
            "has_surveillance": "yes" if str(row[14]).strip() == "1" else "no" if row[14] else None,
            "surveillance_coverage": safe(row[15]),
            "surveillance_type": safe(row[19]),
            "has_formal_plan": "yes" if str(row[20]).strip() == "1" else "no" if row[20] else None,
            "uses_digital_tool": "yes" if str(row[21]).strip() == "1" else "no" if row[21] else None,
            "tool_name": safe(row[22]),
            "tool_url": safe(row[23]),
            "tool_developer": safe(row[24]),
            "tool_date": safe(row[25]),
            "tool_coverage": safe(row[26]),
            "interoperability": safe(row[27]),
            "interop_platforms": safe(row[28]),
            "data_entry_method": safe(row[33]),
            "data_collected": safe(row[36]),
            "trained_staff": "yes" if str(row[41]).strip() == "1" else "no" if row[41] else None,
            "available_all_zones": "yes" if str(row[42]).strip() == "1" else "no" if row[42] else None,
            "limitations": safe(row[43]),
            "main_challenges": safe(row[44]),
            "support_needed": safe(row[45]),
        }
        # Remove None values
        data = {k: v for k, v in data.items() if v is not None}
        submitted_at = safe(row[2]) or "2025-12-01"
        sub_id = deterministic_uuid(CAMPAIGN_A, respondent_id)
        submissions.append((sub_id, CAMPAIGN_A, TEMPLATE_A, data, submitted_at, respondent_id))

    return submissions


def parse_survey_b():
    """Parse Diagnostic Tests survey (41 cols, 21 responses)."""
    from openpyxl import load_workbook
    path = os.path.join(PPR_DIR, "CopyResponses- Information form on diagnostic serological tests for the detection of PPR antibodies used..xlsx")
    wb = load_workbook(path, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    submissions = []
    for row in rows[2:]:
        if not row[0]:
            continue
        respondent_id = str(row[0]).strip()
        data = {
            "country": safe(row[10]) or "",
            "ppr_status": safe(row[11]),
            "last_outbreak": safe(row[13]),
            "pmat_stage": safe(row[14]),
            "notification_system": safe(row[16]),
            "institution_contact": safe(row[18]),
            "contact_email_phone": safe(row[19]),
            "diagnostic_tests": safe(row[20]),
            "kits_requested": safe(row[24]),
            "request_reasons": safe(row[25]),
            "intended_use_date": safe(row[32]),
            "storage_conditions": safe(row[33]),
            "delivery_address": safe(row[35]),
            "kit_receiver": safe(row[36]),
            "official_representative": safe(row[37]),
            "rep_email_phone": safe(row[38]),
            "endorsement": "yes" if row[39] and "Yes" in str(row[39]) else "no" if row[39] else None,
            "commitment": "yes" if row[40] and "Yes" in str(row[40]) else "no" if row[40] else None,
        }
        data = {k: v for k, v in data.items() if v is not None}
        submitted_at = safe(row[2]) or "2025-12-01"
        sub_id = deterministic_uuid(CAMPAIGN_B, respondent_id)
        submissions.append((sub_id, CAMPAIGN_B, TEMPLATE_B, data, submitted_at, respondent_id))

    return submissions


def parse_dataset_c():
    """Parse Kit Allocation ODT (21 shipments)."""
    odt_path = os.path.join(PPR_DIR, "Total PPR kits sent to AUMS 01.04.2025 to 11.05.2026.odt")
    ns = {
        'text': 'urn:oasis:names:tc:opendocument:xmlns:text:1.0',
        'table': 'urn:oasis:names:tc:opendocument:xmlns:table:1.0',
    }
    with zipfile.ZipFile(odt_path) as z:
        with z.open("content.xml") as f:
            tree = ElementTree.parse(f)

    tables = tree.findall('.//table:table', ns)
    if not tables:
        return []

    table = tables[0]
    rows_el = table.findall('.//table:table-row', ns)

    submissions = []
    row_num = 0
    for row_el in rows_el:
        cells = row_el.findall('table:table-cell', ns)
        cell_texts = []
        for cell in cells:
            paras = cell.findall('.//text:p', ns)
            text = ' '.join(p.text or '' for p in paras if p.text).strip()
            cell_texts.append(text)

        # Skip headers and totals
        if not cell_texts or not cell_texts[0].isdigit():
            continue

        row_num += 1
        num = cell_texts[0] if len(cell_texts) > 0 else ""
        date = cell_texts[1] if len(cell_texts) > 1 else ""
        reagent = cell_texts[2] if len(cell_texts) > 2 else "PPR kit"
        kit_format = cell_texts[3] if len(cell_texts) > 3 else ""
        quantity = cell_texts[4] if len(cell_texts) > 4 else ""
        num_samples = cell_texts[5] if len(cell_texts) > 5 else ""
        country = cell_texts[6] if len(cell_texts) > 6 else ""
        year = cell_texts[7] if len(cell_texts) > 7 else ""

        if not country:
            continue

        data = {
            "shipment_date": date,
            "reagent_type": "ppr_kit",
            "kit_format": kit_format,
            "quantity": quantity.replace(",", ""),
            "num_samples": num_samples.replace(",", ""),
            "destination_country": country,
            "year": year or (date.split("/")[-1] if "/" in date else "2025"),
        }
        data = {k: v for k, v in data.items() if v}

        row_key = f"{date}_{country}_{num}"
        sub_id = deterministic_uuid(CAMPAIGN_C, row_key)
        submitted_at = f"20{date.split('/')[-1]}-{date.split('/')[0].zfill(2)}-{date.split('/')[1].zfill(2)}" if "/" in date and len(date.split("/")) == 3 else "2025-06-01"
        # Fix date format (M/D/YYYY → YYYY-MM-DD)
        try:
            parts = date.split("/")
            if len(parts) == 3:
                submitted_at = f"{parts[2]}-{parts[0].zfill(2)}-{parts[1].zfill(2)}"
        except:
            pass

        submissions.append((sub_id, CAMPAIGN_C, TEMPLATE_C, data, submitted_at, row_key))

    return submissions


# ═══════════════════════════════════════════════
# IMPORT TO DB
# ═══════════════════════════════════════════════

def import_submissions(submissions, label):
    """Import submissions via direct SQL on PROD."""
    if not submissions:
        print(f"  [{label}] No data to import")
        return

    # Build SQL
    sql_lines = ["BEGIN;"]
    for sub_id, campaign_id, template_id, data, submitted_at, _ in submissions:
        data_json = json.dumps(data, ensure_ascii=False).replace("'", "''")
        sql_lines.append(
            f"INSERT INTO public.submissions "
            f"(id, tenant_id, campaign_id, template_id, data, submitted_by, submitted_at, "
            f"status, data_classification, version, created_at, updated_at) "
            f"VALUES ('{sub_id}', '{TENANT_AU_IBAR}', '{campaign_id}', '{template_id}', "
            f"'{data_json}'::jsonb, '{USER_SUPER_ADMIN}', '{submitted_at}'::timestamptz, "
            f"'VALIDATED', 'PARTNER', 1, NOW(), NOW()) "
            f"ON CONFLICT (id) DO NOTHING;"
        )
    sql_lines.append("COMMIT;")
    sql_lines.append(f"SELECT count(*) AS total FROM public.submissions WHERE campaign_id = '{submissions[0][1]}';")

    sql = "\n".join(sql_lines)

    # Upload and execute
    c = paramiko.SSHClient()
    c.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    c.connect(VM_APP, username=SSH_USER, password=SSH_PASS, timeout=15)

    tmp = tempfile.NamedTemporaryFile(mode="w", suffix=".sql", delete=False, newline="\n", encoding="utf-8")
    tmp.write(sql)
    tmp.close()

    sftp = c.open_sftp()
    sftp.put(tmp.name, f"/tmp/import_{label}.sql")
    sftp.close()
    os.unlink(tmp.name)

    chan = c.get_transport().open_session()
    chan.exec_command(
        f"echo '{SSH_PASS}' | sudo -S docker run --rm --network host "
        f"-v /tmp/import_{label}.sql:/import.sql:ro "
        f"-e PGPASSWORD={DB_PASS} postgres:16 "
        f"psql -h {DB_HOST} -p 5432 -U aris -d aris -f /import.sql 2>&1 | tail -5"
    )
    chan.settimeout(60)
    out = b""
    try:
        while True:
            ch = chan.recv(4096)
            if not ch: break
            out += ch
    except: pass
    result = out.decode(errors="replace").strip()
    c.close()

    print(f"  [{label}] {len(submissions)} rows → {result}")


def main():
    print("=" * 60)
    print("  PPR SURVEYS — IMPORT EXISTING DATA")
    print("=" * 60)

    # Parse all sources
    print("\n[1/4] Parsing source files...")
    subs_a = parse_survey_a()
    print(f"  Survey A (Surveillance): {len(subs_a)} responses")
    subs_b = parse_survey_b()
    print(f"  Survey B (Diagnostics): {len(subs_b)} responses")
    subs_c = parse_dataset_c()
    print(f"  Dataset C (Kit Allocation): {len(subs_c)} shipments")

    # Import
    print("\n[2/4] Importing Survey A...")
    import_submissions(subs_a, "A")

    print("\n[3/4] Importing Survey B...")
    import_submissions(subs_b, "B")

    print("\n[4/4] Importing Dataset C...")
    import_submissions(subs_c, "C")

    total = len(subs_a) + len(subs_b) + len(subs_c)
    print(f"\n{'=' * 60}")
    print(f"  IMPORT COMPLETE — {total} submissions")
    print(f"{'=' * 60}")


if __name__ == "__main__":
    main()
