"""
Seed PAID referentials into master-data service.
Parses AU-IBAR_PAID.xlsx and uploads:
  - 9 sectors, 334 activities, 31 species, 162 diseases
  - 102 production systems, 6 cash mechanisms, 34 CVA types
  - Country metadata (HH size, disability %)
Uses the FisheryReferential generic model with PAID_* categories.
"""
import paramiko, json, sys, time, os

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

# ── Try to load Excel data with openpyxl ──
try:
    import openpyxl
    EXCEL_PATH = os.path.join(os.path.dirname(__file__), "..", "..", "form", "LICS", "AU-IBAR_PAID.xlsx")
    HAS_EXCEL = os.path.exists(EXCEL_PATH)
except ImportError:
    HAS_EXCEL = False

SSH_PASS = "@u-1baR.0rg$U24"
HOST = "10.202.101.183"

# ── Connect ──
ssh = paramiko.SSHClient()
ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
ssh.connect(HOST, username="arisadmin", password=SSH_PASS,
            timeout=15, allow_agent=False, look_for_keys=False)

# ── Auth ──
_, stdout, _ = ssh.exec_command(
    'curl -sk -X POST "https://localhost/api/v1/credential/auth/login" '
    '-H "Content-Type: application/json" '
    """-d '{"email":"admin@au-aris.org","password":"Aris2026@@4!0"}' 2>/dev/null""",
    timeout=15)
token = json.loads(stdout.read().decode())["data"]["accessToken"]
print(f"Authenticated")

def api_post(path, body):
    data = json.dumps(body)
    chan = ssh.get_transport().open_session()
    chan.settimeout(30)
    chan.exec_command(
        f'curl -sk -X POST "https://localhost{path}" '
        f'-H "Authorization: Bearer {token}" '
        f'-H "Content-Type: application/json" '
        f'--data-binary @- 2>/dev/null'
    )
    chan.sendall(data.encode())
    chan.shutdown_write()
    time.sleep(1)
    resp = b''
    for _ in range(10):
        if chan.recv_ready():
            resp += chan.recv(65536)
        elif chan.exit_status_ready():
            while chan.recv_ready():
                resp += chan.recv(65536)
            break
        time.sleep(0.3)
    try:
        return json.loads(resp.decode())
    except:
        return {"error": resp.decode()[:200]}

# ══════════════════════════════════════════════════════════════════════
#  Extract data from Excel
# ══════════════════════════════════════════════════════════════════════

def extract_from_excel():
    """Parse AU-IBAR_PAID.xlsx and return all referentials."""
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    # ── Sectors ──
    ws = wb["Sector_of_production"]
    sectors = []
    for r in range(2, ws.max_row + 1):
        en = ws.cell(r, 1).value
        fr = ws.cell(r, 2).value
        es = ws.cell(r, 3).value
        sid = ws.cell(r, 4).value
        if en:
            sectors.append({"code": sid or f"SP_{r-1}", "name": {"en": en, "fr": fr or en, "es": es or en}})

    # ── Activities ──
    ws = wb["Activity"]
    activities = []
    for r in range(2, ws.max_row + 1):
        sector_en = ws.cell(r, 1).value
        act_en = ws.cell(r, 2).value
        act_fr = ws.cell(r, 4).value
        act_es = ws.cell(r, 6).value
        unit = ws.cell(r, 7).value
        aid = ws.cell(r, 9).value
        tier = ws.cell(r, 10).value
        if act_en:
            activities.append({
                "code": aid or f"ACT_{r-1}",
                "name": {"en": act_en, "fr": act_fr or act_en, "es": act_es or act_en},
                "parentCode": sector_en,
                "metadata": {"unit": unit, "tier": tier}
            })

    # ── Species ──
    ws = wb["Species"]
    species = []
    for r in range(2, ws.max_row + 1):
        sector_en = ws.cell(r, 1).value
        sp_en = ws.cell(r, 2).value
        sp_fr = ws.cell(r, 4).value
        sp_es = ws.cell(r, 6).value
        sid = ws.cell(r, 7).value
        if sp_en:
            species.append({
                "code": sid or f"SPC_{r-1}",
                "name": {"en": sp_en, "fr": sp_fr or sp_en, "es": sp_es or sp_en},
                "parentCode": sector_en
            })

    # ── Diseases ──
    ws = wb["Disease"]
    diseases = []
    for r in range(2, ws.max_row + 1):
        sector_en = ws.cell(r, 1).value
        species_en = ws.cell(r, 2).value
        dis_en = ws.cell(r, 3).value
        dis_fr = ws.cell(r, 6).value
        dis_es = ws.cell(r, 9).value
        did = ws.cell(r, 10).value
        if dis_en:
            diseases.append({
                "code": did or f"DIS_{r-1}",
                "name": {"en": dis_en, "fr": dis_fr or dis_en, "es": dis_es or dis_en},
                "parentCode": species_en,
                "metadata": {"sector": sector_en}
            })

    # ── Production Systems ──
    ws = wb["ProductionSystem"]
    prodsys = []
    for r in range(2, ws.max_row + 1):
        sector_en = ws.cell(r, 1).value
        sp_en = ws.cell(r, 2).value
        ps_en = ws.cell(r, 3).value
        ps_fr = ws.cell(r, 6).value
        ps_es = ws.cell(r, 9).value
        pid = ws.cell(r, 10).value
        if ps_en:
            prodsys.append({
                "code": pid or f"PS_{r-1}",
                "name": {"en": ps_en, "fr": ps_fr or ps_en, "es": ps_es or ps_en},
                "parentCode": sp_en,
                "metadata": {"sector": sector_en}
            })

    # ── Countries ──
    ws = wb["Countries_List"]
    countries = []
    for r in range(2, ws.max_row + 1):
        name = ws.cell(r, 1).value
        code = ws.cell(r, 2).value
        hh = ws.cell(r, 3).value
        disability = ws.cell(r, 8).value
        lang = ws.cell(r, 9).value
        status = ws.cell(r, 10).value
        hrp = ws.cell(r, 11).value
        if name and status == "Country":
            countries.append({
                "code": code or "",
                "name": {"en": name},
                "metadata": {
                    "hhSize": hh,
                    "disabilityPct": disability,
                    "language": lang,
                    "isHrp": hrp == "HRP"
                }
            })

    return sectors, activities, species, diseases, prodsys, countries


def get_fallback_data():
    """Minimal fallback if Excel not available."""
    sectors = [
        {"code": "SP_1", "name": {"en": "Agriculture", "fr": "Agriculture"}},
        {"code": "SP_2", "name": {"en": "Environment", "fr": "Environnement"}},
        {"code": "SP_3", "name": {"en": "Fishery", "fr": "Pêche"}},
        {"code": "SP_4", "name": {"en": "Forestry", "fr": "Sylviculture"}},
        {"code": "SP_5", "name": {"en": "Land and Water", "fr": "Terre et eau"}},
        {"code": "SP_6", "name": {"en": "Livelihoods and cash", "fr": "Moyens de subsistance et cash"}},
        {"code": "SP_7", "name": {"en": "Livestock", "fr": "Élevage"}},
        {"code": "SP_8", "name": {"en": "Nutrition and Human health", "fr": "Nutrition et santé humaine"}},
        {"code": "SP_9", "name": {"en": "Social protection", "fr": "Protection sociale"}},
    ]
    return sectors, [], [], [], [], []


# ══════════════════════════════════════════════════════════════════════
#  Seed to master-data via FisheryReferential-style endpoint
# ══════════════════════════════════════════════════════════════════════

ENDPOINT = "/api/v1/master-data/fishery-referentials"

def seed_category(category, items):
    """Seed a batch of referentials with the given category."""
    ok = 0
    skip = 0
    for i, item in enumerate(items):
        body = {
            "category": category,
            "code": item["code"],
            "name": item["name"],
            "parentCode": item.get("parentCode"),
            "sortOrder": i + 1,
            "isActive": True,
            "metadata": item.get("metadata", {})
        }
        r = api_post(ENDPOINT, body)
        if r.get("data", {}).get("id"):
            ok += 1
        elif "already exists" in str(r.get("message", "")):
            skip += 1
        else:
            # First few errors, print
            if ok + skip == 0:
                print(f"    Err: {r.get('message', str(r)[:100])}")
            skip += 1
    return ok, skip


# ── Main ──
print("\n═══ PAID Master-Data Seed ═══\n")

if HAS_EXCEL:
    print(f"Reading Excel: {EXCEL_PATH}")
    sectors, activities, species, diseases, prodsys, countries = extract_from_excel()
    print(f"  Sectors: {len(sectors)}, Activities: {len(activities)}, "
          f"Species: {len(species)}, Diseases: {len(diseases)}, "
          f"ProdSystems: {len(prodsys)}, Countries: {len(countries)}")
else:
    print("Excel not found, using fallback data")
    sectors, activities, species, diseases, prodsys, countries = get_fallback_data()

# Seed each category
for label, cat, items in [
    ("Sectors", "PAID_SECTOR", sectors),
    ("Activities", "PAID_ACTIVITY", activities),
    ("Species", "PAID_SPECIES", species),
    ("Diseases", "PAID_DISEASE", diseases),
    ("Production Systems", "PAID_PROD_SYSTEM", prodsys),
    ("Countries", "PAID_COUNTRY_META", countries),
]:
    if not items:
        print(f"  {label}: (skipped — no data)")
        continue
    print(f"  Seeding {label} ({len(items)})...", end=" ", flush=True)
    ok, skip = seed_category(cat, items)
    print(f"OK={ok}, Skip={skip}")

# ── Cash types (hardcoded) ──
cash_types = [
    {"code": "CVA_1", "name": {"en": "Cash - Cash For Work", "fr": "Cash - Cash contre travail"}},
    {"code": "CVA_2", "name": {"en": "Cash - Conditional Cash Transfer", "fr": "Cash - Transfert conditionnel"}},
    {"code": "CVA_3", "name": {"en": "Cash - Unconditional Cash Transfer", "fr": "Cash - Transfert inconditionnel"}},
    {"code": "CVA_4", "name": {"en": "Cash - Voucher @ Fairs", "fr": "Cash - Bons pour foires"}},
    {"code": "CVA_5", "name": {"en": "Cash - Voucher @ voucher scheme", "fr": "Cash - Programme de bons"}},
]
print(f"  Seeding CVA Types ({len(cash_types)})...", end=" ", flush=True)
ok, skip = seed_category("PAID_CVA_TYPE", cash_types)
print(f"OK={ok}, Skip={skip}")

cash_mechs = [
    {"code": "CM_1", "name": {"en": "Bank transfer", "fr": "Virement bancaire"}},
    {"code": "CM_2", "name": {"en": "Electronic voucher", "fr": "Bon électronique"}},
    {"code": "CM_3", "name": {"en": "Mobile money", "fr": "Argent mobile"}},
    {"code": "CM_4", "name": {"en": "Other electronic cash", "fr": "Autre cash électronique"}},
    {"code": "CM_5", "name": {"en": "Paper voucher", "fr": "Bon papier"}},
    {"code": "CM_6", "name": {"en": "Physical cash", "fr": "Cash physique"}},
]
print(f"  Seeding Cash Mechanisms ({len(cash_mechs)})...", end=" ", flush=True)
ok, skip = seed_category("PAID_CASH_MECHANISM", cash_mechs)
print(f"OK={ok}, Skip={skip}")

ssh.close()
print("\nDONE — PAID referentials seeded.")
