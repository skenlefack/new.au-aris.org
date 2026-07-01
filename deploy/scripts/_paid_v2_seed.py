"""
PAID v2 Seed — Parse AMERT Excel (V-03 refine_Nephat) and populate PAID tables.
TRUNCATES all PAID tables then re-inserts 7 LICS projects with full hierarchy.

Usage: python deploy/scripts/_paid_v2_seed.py [--target stg|prod]
"""

import paramiko
import openpyxl
import sys
import os

SSH_USER = "arisadmin"
SSH_PASS = "@u-1baR.0rg$U24"

TARGETS = {
    "stg": {"host": "10.202.101.148", "db_pass": "Ar1s_Stg_2024!xK9mZ", "db_user": "aris"},
    "prod": {"host": "10.202.101.185", "db_pass": "Ar1s_Pr0d_2024!xK9mZ", "db_user": "aris"},
}

EXCEL_PATH = os.path.join(os.path.dirname(__file__), "..", "..",
    "form", "LICS",
    "AMERT SUB- ACTIVITY BREAKDOWN FOR LICS - 2026 - V-03 refine_Nephat.xlsx")

# ── Project metadata ──────────────────────────────────────────
# Maps internal project code → title, type, countries
PROJECT_META = {
    "AQBIOD": {
        "title": "Aquatic Biodiversity (FISH GOV 2)",
        "type": "multiple_countries",
        "countries": ["KE", "TZ", "MZ", "MG", "SN", "GH", "NG", "CM", "MR", "CV"],
    },
    "RAFFS": {
        "title": "Regional Animal Feed and Fodder Systems (RAFFS)",
        "type": "multiple_countries",
        "countries": ["KE", "ET", "NG", "TZ", "UG", "SN", "NE", "ML", "BF", "TD"],
    },
    "ANGR": {
        "title": "Animal Genetics Resources (AnGR)",
        "type": "multiple_countries",
        "countries": ["KE", "ET", "NG", "TZ", "UG", "SN", "CM", "ZA"],
    },
    "LIVESYS": {
        "title": "Climate Resilient and Sustainable Livestock Systems",
        "type": "multiple_countries",
        "countries": ["KE", "ET", "NG", "TZ", "UG", "SN", "NE", "ML", "BF", "TD"],
    },
    "APMD": {
        "title": "African Pastoral Markets Development (APMD)",
        "type": "multiple_countries",
        "countries": ["KE", "ET", "NG", "TZ", "UG", "SN", "NE", "ML", "BF", "TD", "DJ", "SO"],
    },
    "PPPS_RVLC": {
        "title": "Producers-Public-Private Partnerships & Regional Value Chains (PPPPs-RVLC)",
        "type": "multiple_countries",
        "countries": ["KE", "ET", "NG", "TZ", "UG", "SN", "GH", "CM", "ZA", "MZ"],
    },
    "AH_VET_GOV": {
        "title": "Animal Health, One Health, Disease Control & Veterinary Governance",
        "type": "multiple_countries",
        "countries": ["KE", "ET", "NG", "TZ", "UG", "SN", "CM", "GH", "ZA", "MZ"],
    },
}

# ── Excel project name → internal project code ──────────────
PROJECT_NAME_MAP = {
    "AQUATIC BIODIVERSITY": "AQBIOD",
    "FISH GOV 2": "AQBIOD",       # sub-header within AQUATIC BIODIVERSITY
    "RAFFS PROJECT": "RAFFS",
    "ANGR": "ANGR",
    "LIVESYS": "LIVESYS",
    "APMD": "APMD",
    "PPPS - RVLC": "PPPS_RVLC",
    "AH & VET GOV": "AH_VET_GOV",
}

# Executive partners (per project)
EXEC_PARTNERS = {
    "AQBIOD": ["AU-IBAR"],
    "RAFFS": ["AU-IBAR"],
    "ANGR": ["AU-IBAR"],
    "LIVESYS": ["AU-IBAR"],
    "APMD": ["AU-IBAR"],
    "PPPS_RVLC": ["AU-IBAR"],
    "AH_VET_GOV": ["AU-IBAR", "FAO"],
}

# International implementing partners (per project)
INTL_PARTNERS = {
    "AQBIOD": ["FAO", "WorldFish"],
    "RAFFS": ["ILRI", "ICRISAT"],
    "ANGR": ["ILRI", "ICARDA"],
    "LIVESYS": ["ILRI", "FAO", "ICARDA", "CIRAD"],
    "APMD": ["ILRI", "FAO"],
    "PPPS_RVLC": ["ILRI", "FAO"],
    "AH_VET_GOV": ["FAO", "WOAH", "WHO", "ILRI"],
}

# National implementing partners (per project + country)
NATIONAL_PARTNERS = {
    "AQBIOD": {
        "KE": ["Kenya Fisheries Service (KeFS)"],
        "TZ": ["Tanzania Fisheries Research Institute (TAFIRI)"],
        "MZ": ["Instituto Nacional de Investigacao Pesqueira (IIP)"],
        "SN": ["Direction des Peches Maritimes, Senegal"],
        "GH": ["Fisheries Commission, Ghana"],
        "NG": ["Federal Department of Fisheries, Nigeria"],
    },
    "RAFFS": {
        "KE": ["KALRO - Kenya Agricultural and Livestock Research Organization"],
        "ET": ["Ethiopian Institute of Agricultural Research (EIAR)"],
        "NG": ["National Animal Production Research Institute (NAPRI)"],
        "TZ": ["Tanzania Livestock Research Institute (TALIRI)"],
        "SN": ["ISRA - Institut Senegalais de Recherches Agricoles"],
    },
    "ANGR": {
        "KE": ["Kenya Agricultural and Livestock Research Organization (KALRO)", "Ministry of Agriculture, Kenya"],
        "ET": ["Ethiopian Institute of Agricultural Research (EIAR)"],
        "NG": ["National Animal Production Research Institute (NAPRI)"],
        "TZ": ["Tanzania Livestock Research Institute (TALIRI)"],
        "UG": ["National Animal Genetic Resources Centre (NAGRC)"],
        "SN": ["ISRA - Institut Senegalais de Recherches Agricoles"],
        "CM": ["IRAD - Institut de Recherche Agricole pour le Developpement"],
        "ZA": ["Agricultural Research Council (ARC)"],
    },
    "LIVESYS": {
        "KE": ["Ministry of Agriculture, Kenya", "KALRO"],
        "ET": ["Ministry of Agriculture, Ethiopia"],
        "NG": ["Federal Ministry of Agriculture, Nigeria"],
        "TZ": ["Ministry of Livestock and Fisheries, Tanzania"],
        "UG": ["Ministry of Agriculture, Uganda"],
        "SN": ["Ministere de l'Elevage, Senegal"],
    },
    "APMD": {
        "KE": ["Ministry of Agriculture, Kenya"],
        "ET": ["Ministry of Agriculture, Ethiopia"],
        "NG": ["Federal Ministry of Agriculture, Nigeria"],
        "SN": ["Ministere de l'Elevage, Senegal"],
        "DJ": ["Ministere de l'Agriculture, Djibouti"],
    },
    "PPPS_RVLC": {
        "KE": ["Ministry of Agriculture, Kenya"],
        "ET": ["Ministry of Agriculture, Ethiopia"],
        "NG": ["Federal Ministry of Agriculture, Nigeria"],
        "GH": ["Ministry of Food and Agriculture, Ghana"],
    },
    "AH_VET_GOV": {
        "KE": ["DVS Kenya - Directorate of Veterinary Services"],
        "ET": ["Ethiopian Veterinary Drug and Animal Feed Authority"],
        "NG": ["Federal Ministry of Agriculture, Nigeria"],
        "TZ": ["Tanzania Veterinary Laboratory Agency (TVLA)"],
        "GH": ["Veterinary Services Directorate, Ghana"],
        "ZA": ["Department of Agriculture, Land Reform and Rural Development"],
        "SN": ["Direction des Services Veterinaires, Senegal"],
        "ML": ["Direction Nationale des Services Veterinaires, Mali"],
    },
}

# Breakdown fields for training activities
TRAINING_BREAKDOWN = [
    ("n_female_trained", "Number of females trained", "number", 0, False),
    ("n_male_trained", "Number of males trained", "number", 1, False),
]


def esc(s):
    """Escape single quotes for SQL."""
    if s is None:
        return ""
    return str(s).replace("'", "''").strip()


def clean_code(code):
    """Normalize a code: strip whitespace and trailing dots."""
    c = code.strip()
    while c.endswith("."):
        c = c[:-1]
    return c


def parse_excel():
    """Parse the AMERT V-03 Excel and return structured hierarchy.

    Structure:
      Project header → Output (X.X) → Activity (X.X.X.XX) → Sub-activity (X.X.X.XX.XX)
    Logframes are auto-derived from the first 3 segments of Activity codes (X.X.X).
    """
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    ws = wb["Sheet1"]

    current_project = None
    current_output_label = None
    projects = {}

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        col_a = str(row[0] or "").strip()
        col_b = str(row[1] or "").strip()
        col_c = str(row[2] or "").strip()
        col_d = str(row[3] or "").strip()
        col_e = str(row[4] or "").strip()  # sub-activity unit
        col_f = str(row[5] or "").strip()  # PAID activity label
        col_g = str(row[6] or "").strip()  # PAID activity unit

        if not col_a:
            continue

        # ── Project header (no digits at all) ──
        has_digit = any(ch.isdigit() for ch in col_a.replace(" ", ""))
        if not has_digit:
            name_upper = col_a.upper().strip()
            mapped = PROJECT_NAME_MAP.get(name_upper)
            if mapped:
                current_project = mapped
            else:
                # Fallback: use the name as-is
                current_project = name_upper.replace(" ", "_").replace("-", "_")
                print(f"  [WARN] Unmapped project name: '{col_a}' -> {current_project}")

            if current_project not in projects:
                projects[current_project] = {"logframes": {}}
            continue

        if not current_project:
            continue

        code = clean_code(col_a)
        parts = code.split(".")
        num_parts = len(parts)

        # ── Output level: X.X (2 parts) ──
        if num_parts == 2 and col_b:
            current_output_label = col_b
            continue

        # ── Logframe level: X.X.X (3 parts) — explicit row (rare in V-03) ──
        if num_parts == 3 and col_b:
            lf_code = code
            if lf_code not in projects[current_project]["logframes"]:
                projects[current_project]["logframes"][lf_code] = {
                    "label": col_b,
                    "activities": {},
                }
            continue

        # ── Activity level: X.X.X.XX (4 parts) ──
        if num_parts == 4 and col_c:
            # Auto-create logframe from first 3 parts if not already present
            lf_code = ".".join(parts[:3])
            if lf_code not in projects[current_project]["logframes"]:
                # Use the output label or derive from activity
                lf_label = current_output_label or col_c
                projects[current_project]["logframes"][lf_code] = {
                    "label": lf_label,
                    "activities": {},
                }

            projects[current_project]["logframes"][lf_code]["activities"][code] = {
                "label": col_c,
                "subactivities": {},
            }
            continue

        # ── Sub-activity level: X.X.X.XX.XX (5 parts) ──
        if num_parts == 5:
            lf_code = ".".join(parts[:3])
            act_code = ".".join(parts[:4])

            # Build label: prefer col_d, fallback to col_c
            sa_label = col_d or col_c
            if not sa_label:
                continue

            try:
                projects[current_project]["logframes"][lf_code]["activities"][act_code]["subactivities"][code] = {
                    "label": sa_label,
                    "unit": col_e,
                    "paid_activity": col_f,
                    "paid_unit": col_g,
                }
            except KeyError:
                # Activity or logframe missing — auto-create
                if lf_code not in projects[current_project]["logframes"]:
                    projects[current_project]["logframes"][lf_code] = {
                        "label": current_output_label or "Auto-derived",
                        "activities": {},
                    }
                if act_code not in projects[current_project]["logframes"][lf_code]["activities"]:
                    projects[current_project]["logframes"][lf_code]["activities"][act_code] = {
                        "label": f"Activity {act_code}",
                        "subactivities": {},
                    }
                projects[current_project]["logframes"][lf_code]["activities"][act_code]["subactivities"][code] = {
                    "label": sa_label,
                    "unit": col_e,
                    "paid_activity": col_f,
                    "paid_unit": col_g,
                }

    wb.close()
    return projects


def generate_sql(projects):
    """Generate TRUNCATE + INSERT SQL from parsed hierarchy."""
    lines = []

    # ── TRUNCATE all PAID tables (reverse dependency order) ──
    lines.append("-- ============================================================")
    lines.append("-- TRUNCATE all PAID tables")
    lines.append("-- ============================================================")
    lines.append("TRUNCATE TABLE animal_health.paid_breakdown_fields CASCADE;")
    lines.append("TRUNCATE TABLE animal_health.paid_paid_activities CASCADE;")
    lines.append("TRUNCATE TABLE animal_health.paid_subactivities CASCADE;")
    lines.append("TRUNCATE TABLE animal_health.paid_lf_activities CASCADE;")
    lines.append("TRUNCATE TABLE animal_health.paid_logframes CASCADE;")
    lines.append("TRUNCATE TABLE animal_health.paid_impl_partners_national CASCADE;")
    lines.append("TRUNCATE TABLE animal_health.paid_impl_partners_intl CASCADE;")
    lines.append("TRUNCATE TABLE animal_health.paid_executive_partners CASCADE;")
    lines.append("TRUNCATE TABLE animal_health.paid_projects CASCADE;")
    lines.append("")

    # ── Insert projects ──
    lines.append("-- ============================================================")
    lines.append("-- SEED DATA (V-03 refine_Nephat)")
    lines.append("-- ============================================================")
    lines.append("")

    for code, meta in PROJECT_META.items():
        countries_sql = "'{" + ",".join(meta["countries"]) + "}'"
        lines.append(
            f"INSERT INTO animal_health.paid_projects (code, title, type, countries) "
            f"VALUES ('{esc(code)}', '{esc(meta['title'])}', '{meta['type']}', {countries_sql}) "
            f"ON CONFLICT (code) DO NOTHING;"
        )

    lines.append("")

    # ── Executive partners ──
    for proj, partners in EXEC_PARTNERS.items():
        for p in partners:
            lines.append(
                f"INSERT INTO animal_health.paid_executive_partners (project_code, name) "
                f"VALUES ('{proj}', '{esc(p)}');"
            )
    lines.append("")

    # ── International implementing partners ──
    for proj, partners in INTL_PARTNERS.items():
        for p in partners:
            lines.append(
                f"INSERT INTO animal_health.paid_impl_partners_intl (project_code, name) "
                f"VALUES ('{esc(proj)}', '{esc(p)}');"
            )
    lines.append("")

    # ── National implementing partners ──
    for proj, countries in NATIONAL_PARTNERS.items():
        for cc, partners in countries.items():
            for p in partners:
                lines.append(
                    f"INSERT INTO animal_health.paid_impl_partners_national (project_code, country_code, name) "
                    f"VALUES ('{esc(proj)}', '{esc(cc)}', '{esc(p)}');"
                )
    lines.append("")

    # ── Logframes, activities, sub-activities, PAID activities ──
    paid_activity_codes_seen = set()

    for proj_code, proj_data in projects.items():
        if proj_code not in PROJECT_META:
            print(f"  [WARN] Skipping unknown project: {proj_code}")
            continue

        lines.append(f"-- Project: {proj_code}")

        for lf_code, lf_data in proj_data["logframes"].items():
            lines.append(
                f"INSERT INTO animal_health.paid_logframes (code, project_code, label) "
                f"VALUES ('{esc(lf_code)}', '{esc(proj_code)}', '{esc(lf_data['label'][:500])}') "
                f"ON CONFLICT (code) DO NOTHING;"
            )

            for act_code, act_data in lf_data["activities"].items():
                lines.append(
                    f"INSERT INTO animal_health.paid_lf_activities (code, logframe_code, label) "
                    f"VALUES ('{esc(act_code)}', '{esc(lf_code)}', '{esc(act_data['label'][:500])}') "
                    f"ON CONFLICT (code) DO NOTHING;"
                )

                for sa_code, sa_data in act_data["subactivities"].items():
                    lines.append(
                        f"INSERT INTO animal_health.paid_subactivities (code, activity_code, label, unit_of_measure) "
                        f"VALUES ('{esc(sa_code)}', '{esc(act_code)}', '{esc(sa_data['label'][:500])}', '{esc(sa_data['unit'])}') "
                        f"ON CONFLICT (code) DO NOTHING;"
                    )

                    # PAID activity (may be empty)
                    pa_label = sa_data["paid_activity"]
                    pa_unit = sa_data["paid_unit"]
                    if pa_label:
                        pa_code = f"{sa_code}_PA"
                        lines.append(
                            f"INSERT INTO animal_health.paid_paid_activities (code, subactivity_code, label, unit_of_measure) "
                            f"VALUES ('{esc(pa_code)}', '{esc(sa_code)}', '{esc(pa_label)}', '{esc(pa_unit)}') "
                            f"ON CONFLICT (code) DO NOTHING;"
                        )

                        # Add breakdown fields for training activities
                        pa_lower = pa_label.lower()
                        if "training" in pa_lower or "sensitization" in pa_lower or "capacity" in pa_lower:
                            if pa_code not in paid_activity_codes_seen:
                                paid_activity_codes_seen.add(pa_code)
                                for fc, fl, ft, so, req in TRAINING_BREAKDOWN:
                                    lines.append(
                                        f"INSERT INTO animal_health.paid_breakdown_fields "
                                        f"(paid_activity_code, field_code, field_label, field_type, sort_order, is_required) "
                                        f"VALUES ('{esc(pa_code)}', '{fc}', '{fl}', '{ft}', {so}, {str(req).lower()}) "
                                        f"ON CONFLICT (paid_activity_code, field_code) DO NOTHING;"
                                    )

        lines.append("")

    return "\n".join(lines)


def deploy_sql(sql, target_key):
    """Deploy SQL to target database via SSH."""
    t = TARGETS[target_key]
    print(f"\n{'='*60}")
    print(f"  PAID v2 SEED — {target_key.upper()} ({t['host']})")
    print(f"{'='*60}")

    ssh = paramiko.SSHClient()
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    ssh.connect(t["host"], username=SSH_USER, password=SSH_PASS, timeout=15)

    # Write SQL to temp file on server
    sftp = ssh.open_sftp()
    remote_path = "/tmp/_paid_v2_seed.sql"
    with sftp.file(remote_path, "w") as f:
        f.write(sql)
    sftp.close()

    # Determine container name based on target
    pg_container = "aris-stg-postgres" if target_key == "stg" else "aris-postgres"
    sudo = f"echo '{SSH_PASS}' | sudo -S"
    db_url = f"postgresql://{t['db_user']}:{t['db_pass']}@localhost:5432/aris"

    # Copy SQL file into container, then run via psql inside container
    copy_cmd = f"{sudo} docker cp {remote_path} {pg_container}:/tmp/_paid_v2_seed.sql"
    exec_cmd = (
        f"{sudo} docker exec {pg_container} psql "
        f"\"postgresql://{t['db_user']}:{t['db_pass']}@localhost:5432/aris\" "
        f"-f /tmp/_paid_v2_seed.sql"
    )

    print(f"[*] Copying SQL to {pg_container}...")
    ssh.exec_command(copy_cmd, timeout=30)

    print("[*] Running seed SQL...")
    _, stdout, stderr = ssh.exec_command(exec_cmd, timeout=120)
    out = stdout.read().decode("utf-8", errors="replace")
    err = stderr.read().decode("utf-8", errors="replace")

    # Count inserts
    insert_count = out.count("INSERT")
    truncate_count = out.count("TRUNCATE")
    print(f"  Truncated {truncate_count} tables, inserted {insert_count} rows")

    if err.strip():
        err_clean = "\n".join(
            l for l in err.strip().split("\n")
            if "NOTICE" not in l and "password" not in l.lower()
        )
        if err_clean.strip():
            print(f"  Errors:\n{err_clean[:500]}")

    # Verify counts
    verify_sql = """
SELECT 'projects' AS entity, count(*) FROM animal_health.paid_projects
UNION ALL SELECT 'logframes', count(*) FROM animal_health.paid_logframes
UNION ALL SELECT 'activities', count(*) FROM animal_health.paid_lf_activities
UNION ALL SELECT 'subactivities', count(*) FROM animal_health.paid_subactivities
UNION ALL SELECT 'paid_activities', count(*) FROM animal_health.paid_paid_activities
UNION ALL SELECT 'breakdown_fields', count(*) FROM animal_health.paid_breakdown_fields
UNION ALL SELECT 'exec_partners', count(*) FROM animal_health.paid_executive_partners
UNION ALL SELECT 'impl_intl', count(*) FROM animal_health.paid_impl_partners_intl
UNION ALL SELECT 'impl_national', count(*) FROM animal_health.paid_impl_partners_national;
"""
    verify_cmd = (
        f"{sudo} docker exec {pg_container} psql "
        f"\"postgresql://{t['db_user']}:{t['db_pass']}@localhost:5432/aris\" "
        f"-t -c \"{verify_sql}\""
    )
    _, stdout2, _ = ssh.exec_command(verify_cmd, timeout=15)
    verify = stdout2.read().decode("utf-8", errors="replace").strip()
    print(f"\n  Verification:\n{verify}")

    # Cleanup
    ssh.exec_command(f"rm -f {remote_path}")
    ssh.close()
    print(f"\n[OK] Seed complete on {target_key.upper()}")


def main():
    target = "stg"
    if len(sys.argv) > 1:
        arg = sys.argv[1].replace("--target=", "").replace("--target", "").strip()
        if arg in TARGETS:
            target = arg

    print("[*] Parsing AMERT Excel (V-03 refine_Nephat)...")
    projects = parse_excel()

    total_lf = sum(len(p["logframes"]) for p in projects.values())
    total_act = sum(
        len(a["activities"])
        for p in projects.values()
        for a in p["logframes"].values()
    )
    total_sa = sum(
        len(s["subactivities"])
        for p in projects.values()
        for a in p["logframes"].values()
        for s in a["activities"].values()
    )

    print(f"  Projects: {len(projects)}")
    print(f"  LogFrames: {total_lf}")
    print(f"  Activities: {total_act}")
    print(f"  Sub-Activities: {total_sa}")

    for proj, data in projects.items():
        lfs = len(data["logframes"])
        acts = sum(len(a["activities"]) for a in data["logframes"].values())
        sas = sum(
            len(s["subactivities"])
            for a in data["logframes"].values()
            for s in a["activities"].values()
        )
        print(f"    {proj}: {lfs} logframes, {acts} activities, {sas} sub-activities")

    print("\n[*] Generating SQL...")
    sql = generate_sql(projects)
    print(f"  Generated {len(sql)} chars, {sql.count('INSERT')} inserts")

    # Save locally for inspection
    local_path = os.path.join(os.path.dirname(__file__), "_paid_v2_seed_generated.sql")
    with open(local_path, "w", encoding="utf-8") as f:
        f.write(sql)
    print(f"  Saved to {local_path}")

    deploy_sql(sql, target)


if __name__ == "__main__":
    main()
