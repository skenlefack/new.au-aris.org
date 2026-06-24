#!/usr/bin/env python3
"""
ARIS 4.0 — Historical Data Import to Datalake
──────────────────────────────────────────────
Uploads corrected workbooks from mapped/_workbooks/ into the ARIS 4
Historical Data module (datalake service) via the API.

Each workbook becomes a dataset visible at /historical in the web UI.

Usage:
    python deploy/scripts/_import_historical_to_aris4.py
    python deploy/scripts/_import_historical_to_aris4.py --env stg
    python deploy/scripts/_import_historical_to_aris4.py --only "animal_health__Mass_Vaccination"
    python deploy/scripts/_import_historical_to_aris4.py --dry-run
"""
from __future__ import annotations

import argparse
import json
import os
import sys
import time
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

import paramiko

# ── Config ────────────────────────────────────────────────────────────
SSH_USER = os.environ.get("ARIS_DEPLOY_USER", "arisadmin")
SSH_PASS = os.environ.get("ARIS_DEPLOY_PASS", "@u-1baR.0rg$U24")

ENV_CONFIG = {
    "prod": {
        "host": "10.202.101.183",
        "base_url": "https://localhost",
        "container": "aris-datalake",
    },
    "stg": {
        "host": "10.202.101.146",
        "base_url": "https://localhost",
        "container": "aris-stg-datalake",
    },
}

ROOT = Path("C:/new.au-aris.org/DONNEES-HISTORIQUES-ARIS")
WORKBOOKS_DIR = ROOT / "mapped" / "_workbooks"

# ── Dataset metadata per workbook ─────────────────────────────────────
DATASET_META = {
    "animal_health__AU-IBAR_Monthly_Animal_Health_Report__filled": {
        "name": "Monthly Animal Health Reports (2008-2025)",
        "domain": "animal_health",
        "description": "Consolidated disease outbreak reports from 46 African countries. "
                       "Extracted from ARIS 3 historical data (4,989 source files). "
                       "801,318 rows covering disease surveillance, outbreak tracking, "
                       "vaccination status, and control measures.",
        "tags": ["historical", "disease-surveillance", "outbreaks", "2008-2025"],
    },
    "animal_health__Emergency_Disease_Reporting__filled": {
        "name": "Emergency Disease Reports (2008–2025)",
        "domain": "animal_health",
        "description": "Emergency disease notification reports. "
                       "128 records of notifiable disease events across member states.",
        "tags": ["historical", "emergency", "notifications"],
    },
    "animal_health__Mass_Vaccination__filled": {
        "name": "Mass Vaccination Campaigns (2008–2025)",
        "domain": "animal_health",
        "description": "Mass vaccination campaign records. "
                       "375 records covering disease-targeted vaccination activities.",
        "tags": ["historical", "vaccination", "campaigns"],
    },
    "animal_health__Meat_Inspection__filled": {
        "name": "Meat Inspection Reports (2015–2018)",
        "domain": "animal_health",
        "description": "Abattoir meat inspection records. "
                       "992 records with ante-mortem, post-mortem, and lab data.",
        "tags": ["historical", "meat-inspection", "abattoir"],
    },
    "animal_health__Monthly_Vaccination_Report__filled": {
        "name": "Monthly Vaccination Reports (2008–2025)",
        "domain": "animal_health",
        "description": "Monthly vaccination tracking reports. "
                       "6,308 records of routine vaccination activities per species/disease.",
        "tags": ["historical", "vaccination", "monthly"],
    },
    "livestock__Animal_Population_and_Composition__filled": {
        "name": "Animal Population & Composition (2008–2025)",
        "domain": "livestock-prod",
        "description": "Livestock population census data. "
                       "249 records of animal population by species, age group, and production system.",
        "tags": ["historical", "census", "population", "livestock"],
    },
    "trade_sps__Import_and_Export__filled": {
        "name": "Animal Import & Export Records (2014–2017)",
        "domain": "trade_sps",
        "description": "Import/export records of animals and animal products. "
                       "90 records covering trade flows between member states.",
        "tags": ["historical", "trade", "import-export"],
    },
}


def ssh_exec(ssh: paramiko.SSHClient, cmd: str, timeout: int = 60) -> str:
    """Execute command via SSH, return stdout."""
    _, stdout, stderr = ssh.exec_command(cmd, timeout=timeout)
    out = stdout.read().decode("utf-8", errors="replace")
    err = stderr.read().decode("utf-8", errors="replace")
    if err and "warning" not in err.lower():
        print(f"  [stderr] {err[:200]}", flush=True)
    return out


def login(ssh: paramiko.SSHClient, base_url: str) -> str:
    """Login and return access token."""
    body = '{"email":"admin@au-aris.org","password":"Aris2026@@4!0"}'
    cmd = (
        f'curl -sk -X POST "{base_url}/api/v1/credential/auth/login" '
        f'-H "Content-Type: application/json" -d \'{body}\''
    )
    raw = ssh_exec(ssh, cmd)
    resp = json.loads(raw)
    return resp["data"]["accessToken"]


def upload_workbook(
    ssh: paramiko.SSHClient,
    sftp: paramiko.SFTPClient,
    base_url: str,
    token: str,
    local_path: Path,
    meta: dict,
) -> dict | None:
    """Upload a workbook to the historical data import API."""
    file_name = local_path.name

    # 1. Transfer file to server via SFTP
    remote_tmp = f"/tmp/aris_hist_{file_name}"
    print(f"  [sftp] uploading {local_path.name} ({local_path.stat().st_size / 1024 / 1024:.1f} MB)...",
          flush=True)
    sftp.put(str(local_path), remote_tmp)

    # 2. Import via curl multipart from server
    tags_json = json.dumps(meta.get("tags", []))
    cmd = (
        f'curl -sk -X POST "{base_url}/api/v1/historical/import" '
        f'-H "Authorization: Bearer {token}" '
        f'-F "file=@{remote_tmp};filename={file_name}" '
        f'-F "name={meta["name"]}" '
        f'-F "domain={meta["domain"]}" '
        f'-F "description={meta.get("description", "")}" '
        f"-F 'tags={tags_json}' "
        f'--max-time 600'
    )
    print(f"  [api] importing via /api/v1/historical/import...", flush=True)
    raw = ssh_exec(ssh, cmd, timeout=600)

    # 3. Cleanup
    try:
        sftp.remove(remote_tmp)
    except Exception:
        pass

    # 4. Parse response
    try:
        resp = json.loads(raw)
        if "data" in resp:
            ds = resp["data"]
            print(f"  [ok] dataset created: id={ds.get('id','?')}, "
                  f"rows={ds.get('rowCount','?')}, status={ds.get('status','?')}",
                  flush=True)
            return ds
        else:
            print(f"  [err] API response: {raw[:500]}", flush=True)
            return None
    except json.JSONDecodeError:
        print(f"  [err] Non-JSON response: {raw[:500]}", flush=True)
        return None


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--env", default="prod", choices=["prod", "stg"])
    ap.add_argument("--only", default=None,
                    help="Import only this workbook (partial match on filename)")
    ap.add_argument("--dry-run", action="store_true",
                    help="Show what would be imported without actually doing it")
    args = ap.parse_args()

    cfg = ENV_CONFIG[args.env]
    host = cfg["host"]
    base_url = cfg["base_url"]

    # Collect workbooks to import
    workbooks = []
    for wb_path in sorted(WORKBOOKS_DIR.glob("*__filled.xlsx")):
        slug = wb_path.stem
        if args.only and args.only not in slug:
            continue
        meta = DATASET_META.get(slug)
        if not meta:
            print(f"[skip] {slug} — no metadata defined")
            continue
        workbooks.append((wb_path, meta))

    if not workbooks:
        print("[err] No workbooks found to import.")
        return 1

    print(f"[main] Environment: {args.env} ({host})")
    print(f"[main] Workbooks to import: {len(workbooks)}")
    for wb, meta in workbooks:
        size_mb = wb.stat().st_size / 1024 / 1024
        print(f"  - {wb.name} ({size_mb:.1f} MB) → {meta['name']}")
    print()

    if args.dry_run:
        print("[dry-run] Would import the above workbooks. Exiting.")
        return 0

    # Connect
    print(f"[ssh] connecting to {host}...", flush=True)
    ssh = paramiko.SSHClient()
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    ssh.connect(host, username=SSH_USER, password=SSH_PASS,
                timeout=20, allow_agent=False, look_for_keys=False)
    sftp = ssh.open_sftp()

    # Login
    print("[auth] logging in...", flush=True)
    token = login(ssh, base_url)
    print("[auth] ok\n", flush=True)

    # Import each workbook
    results = []
    for i, (wb_path, meta) in enumerate(workbooks, 1):
        slug = wb_path.stem
        print(f"\n[{i}/{len(workbooks)}] === {meta['name']} ===", flush=True)

        # Skip the huge file (>50MB) — API has 50MB limit
        file_size = wb_path.stat().st_size
        if file_size > 100 * 1024 * 1024:
            print(f"  [warn] File too large ({file_size / 1024 / 1024:.1f} MB > 100 MB limit)")
            print(f"  [info] Will split into chunks...", flush=True)
            # Split large workbook into country-based chunks
            chunks = split_workbook(wb_path, meta)
            for chunk_path, chunk_meta in chunks:
                ds = upload_workbook(ssh, sftp, base_url, token, chunk_path, chunk_meta)
                results.append((chunk_meta["name"], ds))
                time.sleep(2)  # Avoid overloading
        else:
            ds = upload_workbook(ssh, sftp, base_url, token, wb_path, meta)
            results.append((meta["name"], ds))

        time.sleep(1)

    sftp.close()
    ssh.close()

    # Summary
    print("\n" + "=" * 70)
    print("IMPORT SUMMARY")
    print("=" * 70)
    ok = 0
    for name, ds in results:
        if ds and ds.get("status") == "READY":
            ok += 1
            print(f"  ✓ {name}: {ds.get('rowCount', '?')} rows")
        else:
            print(f"  ✗ {name}: FAILED")
    print(f"\n{ok}/{len(results)} datasets imported successfully.")
    return 0 if ok == len(results) else 1


def split_workbook(wb_path: Path, meta: dict, chunk_size: int = 80_000) -> list[tuple[Path, dict]]:
    """Split a large workbook into row-count-based chunks."""
    import openpyxl

    print(f"  [split] loading workbook...", flush=True)
    wb = openpyxl.load_workbook(wb_path, read_only=True)
    ws = wb["Data"]
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    headers = all_rows[0]
    labels = all_rows[1]
    types = all_rows[2]
    data_rows = all_rows[3:]
    total = len(data_rows)
    num_chunks = (total + chunk_size - 1) // chunk_size

    print(f"  [split] {total:,} rows → {num_chunks} chunks of ~{chunk_size:,}", flush=True)

    chunk_dir = wb_path.parent / "_chunks"
    chunk_dir.mkdir(exist_ok=True)

    chunks = []
    for ci in range(num_chunks):
        start = ci * chunk_size
        end = min(start + chunk_size, total)
        chunk_rows = data_rows[start:end]

        part = ci + 1
        chunk_name = f"{wb_path.stem}__part{part:02d}.xlsx"
        chunk_path = chunk_dir / chunk_name

        print(f"    Part {part}/{num_chunks}: rows {start+1:,}–{end:,} ({len(chunk_rows):,} rows)...",
              flush=True)

        cwb = openpyxl.Workbook()
        cws = cwb.active
        cws.title = "Data"
        cws.append(list(headers))
        cws.append(list(labels))
        cws.append(list(types))
        for row in chunk_rows:
            cws.append(list(row))
        cwb.save(chunk_path)
        cwb.close()

        base_name = meta["name"].split("—")[1].strip() if "—" in meta["name"] else meta["name"]
        chunk_meta = {
            "name": f"ARIS 3 — {base_name} (Part {part}/{num_chunks})",
            "domain": meta["domain"],
            "description": f"{meta['description']} Part {part}/{num_chunks}, rows {start+1:,}–{end:,}.",
            "tags": meta.get("tags", []) + [f"part-{part}"],
        }
        chunks.append((chunk_path, chunk_meta))

    return chunks


if __name__ == "__main__":
    sys.exit(main())
